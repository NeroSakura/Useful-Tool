import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 询问数据源文件夹路径
while True:
    folder_path = input("请输入数据源文件夹的完整路径：")
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
        # 获取文件夹内所有 Excel 文件
        excel_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith(('.xlsx', '.xls'))]
        if not excel_files:
            print("该文件夹内没有找到 Excel 文件，请重新输入。")
        else:
            break
    else:
        print("输入的路径不是有效的文件夹，请重新输入。")

# 获取桌面路径
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

# 定义要保留的列
selected_columns = ['线索3级来源', '地区', '大区', 'SQL#', 'SQL $M', '商机 $M', '订单 $M', 'SQL $M 差额', '商机 $M 差额', '订单 $M 差额']

# 定义要新增的列
new_columns = ['高价值客户覆盖数', 'SQL目标', '订单目标', '高价值客户覆盖目标', 'SQL达成率', '订单转化率', '订单达成率']

try:
    # 获取当前日期
    now = datetime.now()
    date_str = now.strftime("%m月%d日")

    # 生成保存文件的完整路径
    save_file_path = os.path.join(desktop_path, f'比对结果{date_str}.xlsx')

    # 创建 ExcelWriter 对象
    with pd.ExcelWriter(save_file_path, engine='openpyxl') as writer:
        for file_path in excel_files:
            # 读取 Excel 文件
            excel_file = pd.ExcelFile(file_path)
            # 获取所有表名
            sheet_names = excel_file.sheet_names

            for sheet_name in sheet_names:
                df = excel_file.parse(sheet_name)

                # 保留指定列
                df = df[selected_columns]

                # 新增列，初始化为空
                for col in new_columns:
                    df[col] = None

                # 计算 SQL 达成率
                def calculate_sql_rate(row):
                    if pd.notna(row['SQL $M']) and pd.notna(row['SQL目标']) and row['SQL目标'] != 0:
                        return row['SQL $M'] / row['SQL目标']
                    return None

                df['SQL达成率'] = df.apply(calculate_sql_rate, axis=1)

                # 计算订单转化率
                def calculate_order_conversion(row):
                    if pd.notna(row['订单 $M']) and pd.notna(row['SQL $M']) and row['SQL $M'] != 0:
                        return row['订单 $M'] / row['SQL $M']
                    return None

                df['订单转化率'] = df.apply(calculate_order_conversion, axis=1)

                # 计算订单达成率
                def calculate_order_achievement(row):
                    if pd.notna(row['订单 $M']) and pd.notna(row['订单目标']) and row['订单目标'] != 0:
                        return row['订单 $M'] / row['订单目标']
                    return None

                df['订单达成率'] = df.apply(calculate_order_achievement, axis=1)

                # 调整列顺序，将 SQL达成率、订单转化率、订单达成率放到 H、I、J 列
                columns = df.columns.tolist()
                columns.remove('SQL达成率')
                columns.remove('订单转化率')
                columns.remove('订单达成率')
                new_columns_order = columns[:7] + ['SQL达成率', '订单转化率', '订单达成率'] + columns[7:]
                df = df[new_columns_order]

                # 将修改后的 DataFrame 写回 Excel 文件
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 加载保存的文件进行格式设置
    wb = load_workbook(save_file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置整体字体、字号和对齐方式
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置首行格式
        for cell in ws[1]:
            cell.font = Font(name='微软雅黑', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 35

        # 设置其余行高
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 22.5

        # 冻结窗格
        ws.freeze_panes = ws['A2']

        # 设置列宽
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['D'].width = 7.5
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 8
        for col in range(2, 11):
            if col != 4:
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 10
        for col in range(11, 18):
            if col not in [12, 13]:
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 11

        # 设置百分比格式
        # 由于列顺序调整，更新列索引
        sql_rate_col_index = 8
        order_conv_col_index = 9
        order_ach_col_index = 10
        for col_index in [sql_rate_col_index, order_conv_col_index, order_ach_col_index]:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_index)
                if cell.value is not None:
                    cell.number_format = '0.00%'

        # 设置末行汇总行格式
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(name='微软雅黑', size=11, bold=True)
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    # 确保至少有一个工作表是可见的
    for sheet in wb:
        sheet.sheet_state = 'visible'

    # 保存修改后的文件
    wb.save(save_file_path)
    print("文件处理完成，已保存到桌面。")

except PermissionError:
    print("没有足够的权限访问该文件，请检查文件权限。")
except Exception as e:
    print(f"发生未知错误：{e}")
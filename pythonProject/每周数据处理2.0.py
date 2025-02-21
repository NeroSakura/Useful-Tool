import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re

# 1. 向用户询问包含上周和本周数据文件的文件夹路径
folder_path = input("请输入包含上周和本周数据文件的文件夹路径：")

# 检查文件夹路径是否存在
if not os.path.exists(folder_path):
    print("输入的文件夹路径不存在，请检查后重新运行。")
    exit(1)

# 获取文件夹内所有的 .xlsx 文件
xlsx_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith('.xlsx')]

# 根据文件名中的日期信息排序，假设文件名中包含日期（如 20250207、20250215）
def get_date_from_filename(filename):
    match = re.search(r'\d{8}', filename)
    return int(match.group()) if match else 0

xlsx_files.sort(key=get_date_from_filename)

# 确保至少有两个文件
if len(xlsx_files) < 2:
    print("文件夹中至少需要有两个 .xlsx 文件，请检查后重新运行。")
    exit(1)

# 假设最后两个文件分别是本周和上周的数据文件
last_week_file = xlsx_files[-2]
this_week_file = xlsx_files[-1]

# 获取用户桌面路径
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
output_file = os.path.join(desktop_path, "比对结果.xlsx")

# 定义 sheet 页名称和需要比对的列
sheet_names = ["智慧中国行", "客户研讨会", "AI科技品鉴会", "创新之旅"]
compare_columns = ["SQL $M", "商机 $M", "订单 $M"]
new_column_names = ["SQL $M 差额", "商机 $M 差额", "订单 $M 差额"]

# 创建一个 ExcelWriter 对象
try:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name in sheet_names:
            try:
                # 读取上周和本周的数据
                last_week_df = pd.read_excel(last_week_file, sheet_name=sheet_name)
                this_week_df = pd.read_excel(this_week_file, sheet_name=sheet_name)

                # 确保参与计算的列是数值类型
                for col in compare_columns:
                    last_week_df[col] = pd.to_numeric(last_week_df[col], errors='coerce')
                    this_week_df[col] = pd.to_numeric(this_week_df[col], errors='coerce')

                # 合并两表数据，以本周数据为主
                merged_df = pd.merge(this_week_df, last_week_df, on=["线索3级来源", "cdbid", "leadsid 的计数", "IQL#", "MQL#", "MQLPro#", "SQL#"], how="left", suffixes=('', '_last'))

                # 计算比对列的差额
                for col, new_col in zip(compare_columns, new_column_names):
                    last_col = col + "_last"
                    merged_df[new_col] = merged_df[col] - merged_df[last_col]

                # 去掉合并时产生的后缀列
                for col in compare_columns:
                    last_col = col + "_last"
                    if last_col in merged_df.columns:
                        merged_df.drop(columns=[last_col], inplace=True)

                # 找出新增的行项目（以线索3级来源为基准）
                last_week_sources = set(last_week_df["线索3级来源"])
                new_rows = merged_df[~merged_df["线索3级来源"].isin(last_week_sources)].index

                # 标记新增行
                merged_df["is_new"] = merged_df.index.isin(new_rows)

                # 在 A 列后新增“地区”列
                merged_df.insert(1, '地区', '')

                # 从 A 列提取地区信息填充到“地区”列
                def extract_region(text):
                    if isinstance(text, str):
                        # 先将 丨 和 ｜ 都替换成  |
                        text = text.replace("丨", " | ").replace("｜", " | ")
                        parts = text.split("|")
                        if len(parts) > 1:
                            return parts[-1].strip()
                    return ''

                merged_df['地区'] = merged_df['线索3级来源'].apply(extract_region)

                # 将结果写入 Excel
                merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                print(f"处理工作表 {sheet_name} 时出现错误: {e}")

    # 加载生成的 Excel 文件进行样式设置
    wb = load_workbook(output_file)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        max_row = ws.max_row
        header = [cell.value for cell in ws[1]]

        # 找到新列的列索引
        new_col_indices = [header.index(col) + 1 for col in new_column_names]

        # 找到 is_new 列的索引
        is_new_col_index = header.index("is_new") + 1

        # 设置字体、列宽和行高
        default_font = Font(name='微软雅黑', size=11)
        bold_font = Font(name='微软雅黑', size=11, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                if row[0].row == 1:
                    cell.font = bold_font
                else:
                    cell.font = default_font
                cell.alignment = center_alignment
        # 设置列宽
        ws.column_dimensions['A'].width = 29
        ws.column_dimensions['B'].width = 12  # 新增“地区”列的列宽
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions[chr(ord('A') + 10)].width = 13.5
        ws.column_dimensions[chr(ord('A') + 11)].width = 13.5
        ws.column_dimensions[chr(ord('A') + 12)].width = 13.5

        for row in ws.iter_rows(min_row=1, max_row=max_row):
            ws.row_dimensions[row[0].row].height = 18

        # 遍历每一行，设置样式并进行四舍五入
        for row in range(2, max_row + 1):
            # 检查是否为新增行
            is_new_row = ws.cell(row=row, column=is_new_col_index).value
            if is_new_row:
                for col in range(1, ws.max_column):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # 设置差额列的样式并进行四舍五入
            for col_index in new_col_indices:
                value = ws.cell(row=row, column=col_index).value
                if pd.notna(value):
                    rounded_value = round(value, 2) if abs(value) >= 0.01 else None
                    ws.cell(row=row, column=col_index).value = rounded_value
                    if rounded_value and rounded_value > 0:
                        ws.cell(row=row, column=col_index).font = Font(name='微软雅黑', size=11, bold=True)
                    elif rounded_value and rounded_value < 0:
                        ws.cell(row=row, column=col_index).font = Font(name='微软雅黑', size=11, color="FF0000", bold=True)

        # 删除 is_new 列
        ws.delete_cols(is_new_col_index)

        # 增加汇总行
        summary_row = max_row + 1
        ws.cell(row=summary_row, column=1, value='汇总')
        ws.cell(row=summary_row, column=1).font = bold_font
        for col_index, col_name in enumerate(header, start=1):
            if col_name in compare_columns + new_column_names:
                column_values = [ws.cell(row=row, column=col_index).value for row in range(2, max_row + 1) if
                                 pd.notna(ws.cell(row=row, column=col_index).value)]
                summary_value = sum(column_values)
                rounded_summary = round(summary_value, 2) if abs(summary_value) >= 0.01 else None
                ws.cell(row=summary_row, column=col_index, value=rounded_summary)
                if rounded_summary and rounded_summary > 0:
                    ws.cell(row=summary_row, column=col_index).font = Font(name='微软雅黑', size=11, bold=True)
                elif rounded_summary and rounded_summary < 0:
                    ws.cell(row=summary_row, column=col_index).font = Font(name='微软雅黑', size=11, color="FF0000",
                                                                           bold=True)

        # 冻结首行
        ws.freeze_panes = ws['A2']

        # 设置汇总行样式
        summary_fill = PatternFill(start_color='62B460', end_color='62B460', fill_type='solid')
        for col in range(1, len(header) + 1):
            ws.cell(row=summary_row, column=col).fill = summary_fill
            ws.cell(row=summary_row, column=col).alignment = center_alignment

    # 保存修改后的 Excel 文件
    wb.save(output_file)
    print(f"比对结果已保存至 {output_file}")
except Exception as e:
    print(f"处理过程中出现错误: {e}")
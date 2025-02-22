import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re

# 优化后的区域映射表（基于行政划分）
region_mapping = {
    "东南大区": {"覆盖": ["浙江", "江西", "福建"]},
    "华东大区": {"覆盖": ["上海", "江苏", "安徽"]},
    "华北大区": {"覆盖": ["北京", "天津", "河北", "山西", "内蒙古"]},
    "华南大区": {"覆盖": ["广东", "广西", "海南"]},
    "西南大区": {"覆盖": ["四川", "重庆", "云南", "贵州", "西藏"]},
    "西北大区": {"覆盖": ["陕西", "甘肃", "宁夏", "青海", "新疆"]},
    "中东大区": {"覆盖": ["山东", "河南", "湖北", "湖南"]},
    "东北大区": {"覆盖": ["黑龙江", "吉林", "辽宁"]}
}

# 使用标准城市-省份映射（完整数据）
city_to_province = {
    **{city: "浙江" for city in ["杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华", "衢州", "舟山", "台州", "丽水"]},
    **{city: "江苏" for city in ["南京", "苏州", "无锡", "常州", "徐州", "南通", "连云港", "淮安", "盐城", "扬州", "镇江", "泰州", "宿迁"]},
    **{city: "安徽" for city in ["合肥", "芜湖", "蚌埠", "淮南", "马鞍山", "淮北", "铜陵", "安庆", "黄山", "阜阳", "宿州", "滁州", "六安", "宣城", "池州", "亳州"]},
    **{city: "江西" for city in ["南昌", "九江", "景德镇", "萍乡", "新余", "鹰潭", "赣州", "吉安", "宜春", "抚州", "上饶"]},
    **{city: "福建" for city in ["福州", "厦门", "莆田", "三明", "泉州", "漳州", "南平", "龙岩", "宁德"]},
    **{city: "上海" for city in ["上海"]},
    **{city: "北京" for city in ["北京"]},
    **{city: "天津" for city in ["天津"]},
    **{city: "河北" for city in ["石家庄", "唐山", "秦皇岛", "邯郸", "邢台", "保定", "张家口", "承德", "沧州", "廊坊", "衡水"]},
    **{city: "山西" for city in ["太原", "大同", "阳泉", "长治", "晋城", "朔州", "晋中", "运城", "忻州", "临汾", "吕梁"]},
    **{city: "内蒙古" for city in ["呼和浩特", "包头", "乌海", "赤峰", "通辽", "鄂尔多斯", "呼伦贝尔", "巴彦淖尔", "乌兰察布", "兴安盟", "锡林郭勒盟", "阿拉善盟"]},
    **{city: "广东" for city in ["广州", "深圳", "珠海", "汕头", "佛山", "韶关", "湛江", "肇庆", "江门", "茂名", "惠州", "梅州", "汕尾", "河源", "阳江", "清远", "东莞", "中山", "潮州", "揭阳", "云浮"]},
    **{city: "广西" for city in ["南宁", "柳州", "桂林", "梧州", "北海", "防城港", "钦州", "贵港", "玉林", "百色", "贺州", "河池", "来宾", "崇左"]},
    **{city: "海南" for city in ["海口", "三亚", "三沙", "儋州"]},
    **{city: "四川" for city in ["成都", "自贡", "攀枝花", "泸州", "德阳", "绵阳", "广元", "遂宁", "内江", "乐山", "南充", "眉山", "宜宾", "广安", "达州", "雅安", "巴中", "资阳", "阿坝藏族羌族自治州", "甘孜藏族自治州", "凉山彝族自治州"]},
    **{city: "重庆" for city in ["重庆"]},
    **{city: "云南" for city in ["昆明", "曲靖", "玉溪", "保山", "昭通", "丽江", "普洱", "临沧", "楚雄彝族自治州", "红河哈尼族彝族自治州", "文山壮族苗族自治州", "西双版纳傣族自治州", "大理白族自治州", "德宏傣族景颇族自治州", "怒江傈僳族自治州", "迪庆藏族自治州"]},
    **{city: "贵州" for city in ["贵阳", "六盘水", "遵义", "安顺", "毕节", "铜仁", "黔西南布依族苗族自治州", "黔东南苗族侗族自治州", "黔南布依族苗族自治州"]},
    **{city: "西藏" for city in ["拉萨", "日喀则", "昌都", "林芝", "山南", "那曲", "阿里地区"]},
    **{city: "陕西" for city in ["西安", "铜川", "宝鸡", "咸阳", "渭南", "延安", "汉中", "榆林", "安康", "商洛"]},
    **{city: "甘肃" for city in ["兰州", "嘉峪关", "金昌", "白银", "天水", "武威", "张掖", "平凉", "酒泉", "庆阳", "定西", "陇南", "临夏回族自治州", "甘南藏族自治州"]},
    **{city: "宁夏" for city in ["银川", "石嘴山", "吴忠", "固原", "中卫"]},
    **{city: "青海" for city in ["西宁", "海东", "海北藏族自治州", "黄南藏族自治州", "海南藏族自治州", "果洛藏族自治州", "玉树藏族自治州", "海西蒙古族藏族自治州"]},
    **{city: "新疆" for city in ["乌鲁木齐", "克拉玛依", "吐鲁番", "哈密", "昌吉回族自治州", "博尔塔拉蒙古自治州", "巴音郭楞蒙古自治州", "阿克苏地区", "克孜勒苏柯尔克孜自治州", "喀什", "喀什地区", "和田地区", "伊犁哈萨克自治州", "塔城地区", "阿勒泰地区"]},
    **{city: "山东" for city in ["济南", "青岛", "淄博", "枣庄", "东营", "烟台", "潍坊", "济宁", "泰安", "威海", "日照", "临沂", "德州", "聊城", "滨州", "菏泽"]},
    **{city: "河南" for city in ["郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁", "新乡", "焦作", "濮阳", "许昌", "漯河", "三门峡", "南阳", "商丘", "信阳", "周口", "驻马店"]},
    **{city: "湖北" for city in ["武汉", "黄石", "十堰", "宜昌", "襄阳", "鄂州", "荆门", "孝感", "荆州", "黄冈", "咸宁", "随州", "恩施土家族苗族自治州"]},
    **{city: "湖南" for city in ["长沙", "株洲", "湘潭", "衡阳", "邵阳", "岳阳", "常德", "张家界", "益阳", "郴州", "永州", "怀化", "娄底", "湘西土家族苗族自治州"]},
    **{city: "黑龙江" for city in ["哈尔滨", "齐齐哈尔", "鸡西", "鹤岗", "双鸭山", "大庆", "伊春", "佳木斯", "七台河", "牡丹江", "黑河", "绥化", "大兴安岭地区"]},
    **{city: "吉林" for city in ["长春", "吉林", "四平", "辽源", "通化", "白山", "松原", "白城", "延边朝鲜族自治州"]},
    **{city: "辽宁" for city in ["沈阳", "大连", "鞍山", "抚顺", "本溪", "丹东", "锦州", "营口", "阜新", "辽阳", "盘锦", "铁岭", "朝阳", "葫芦岛"]}
}


def match_region(area):
    # 优先通过城市找省份
    province = next((p for city, p in city_to_province.items() if city in area), None)
    if province:
        return next((k for k, v in region_mapping.items() if province in v["覆盖"]), '')

    # 次优先直接匹配省份
    return next((k for k, v in region_mapping.items() if any(p in area for p in v["覆盖"])), '')


# 1. 向用户询问包含上周和本周数据文件的文件夹路径
folder_path = input("请输入包含上周和本周数据文件的文件夹路径：")

# 检查文件夹路径是否存在
if not os.path.exists(folder_path):
    print("输入的文件夹路径不存在，请检查后重新运行。")
    exit(1)

# 获取文件夹内所有的 .xlsx 文件
xlsx_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith('.xlsx')]

# 根据文件名中的日期信息排序，假设文件名中包含日期（如 20250207、20250215）
def get_date_from_filename(filename):
    match = re.search(r'\d{8}', filename)
    return int(match.group()) if match else 0

xlsx_files.sort(key=get_date_from_filename)

# 确保至少有两个文件
if len(xlsx_files) < 2:
    print("文件夹中至少需要有两个 .xlsx 文件，请检查后重新运行。")
    exit(1)

# 假设最后两个文件分别是本周和上周的数据文件
last_week_file = xlsx_files[-2]
this_week_file = xlsx_files[-1]

# 获取用户桌面路径
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
output_file = os.path.join(desktop_path, "比对结果.xlsx")

# 定义 sheet 页名称和需要比对的列
sheet_names = ["智慧中国行", "客户研讨会", "AI科技品鉴会", "创新之旅"]
compare_columns = ["SQL $M", "商机 $M", "订单 $M"]
new_column_names = ["SQL $M 差额", "商机 $M 差额", "订单 $M 差额"]

# 创建一个 ExcelWriter 对象
try:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name in sheet_names:
            try:
                # 读取上周和本周的数据
                last_week_df = pd.read_excel(last_week_file, sheet_name=sheet_name)
                this_week_df = pd.read_excel(this_week_file, sheet_name=sheet_name)

                # 确保参与计算的列是数值类型
                for col in compare_columns:
                    last_week_df[col] = pd.to_numeric(last_week_df[col], errors='coerce')
                    this_week_df[col] = pd.to_numeric(this_week_df[col], errors='coerce')

                # 合并两表数据，以本周数据为主
                merged_df = pd.merge(this_week_df, last_week_df, on=["线索3级来源", "cdbid", "leadsid 的计数", "IQL#", "MQL#", "MQLPro#", "SQL#"], how="left", suffixes=('', '_last'))

                # 计算比对列的差额
                for col, new_col in zip(compare_columns, new_column_names):
                    last_col = col + "_last"
                    merged_df[new_col] = merged_df[col] - merged_df[last_col]

                # 去掉合并时产生的后缀列
                for col in compare_columns:
                    last_col = col + "_last"
                    if last_col in merged_df.columns:
                        merged_df.drop(columns=[last_col], inplace=True)

                # 找出新增的行项目（以线索3级来源为基准）
                last_week_sources = set(last_week_df["线索3级来源"])
                new_rows = merged_df[~merged_df["线索3级来源"].isin(last_week_sources)].index

                # 标记新增行
                merged_df["is_new"] = merged_df.index.isin(new_rows)

                # 在 A 列后新增“地区”列
                merged_df.insert(1, '地区', '')

                # 从 A 列提取地区信息填充到“地区”列
                def extract_region(text):
                    if isinstance(text, str):
                        # 先将 丨 和 ｜ 都替换成  |
                        text = text.replace("丨", " | ").replace("｜", " | ")
                        parts = text.split("|")
                        if len(parts) > 1:
                            return parts[-1].strip()
                    return ''

                merged_df['地区'] = merged_df['线索3级来源'].apply(extract_region)

                # 在 B 列后新增“大区”列
                merged_df.insert(2, '大区', '')

                # 根据“地区”列匹配大区
                merged_df['大区'] = merged_df['地区'].apply(match_region)

                # 将结果写入 Excel
                merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                print(f"处理工作表 {sheet_name} 时出现错误: {e}")

    # 加载生成的 Excel 文件进行样式设置
    wb = load_workbook(output_file)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        max_row = ws.max_row
        header = [cell.value for cell in ws[1]]

        # 找到新列的列索引
        new_col_indices = [header.index(col) + 1 for col in new_column_names]

        # 找到 is_new 列的索引
        is_new_col_index = header.index("is_new") + 1

        # 设置字体、列宽和行高
        default_font = Font(name='微软雅黑', size=11)
        bold_font = Font(name='微软雅黑', size=11, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                if row[0].row == 1:
                    cell.font = bold_font
                else:
                    cell.font = default_font
                cell.alignment = center_alignment
        # 设置列宽
        ws.column_dimensions['A'].width = 31
        ws.column_dimensions['B'].width = 12  # 新增“地区”列的列宽
        ws.column_dimensions['C'].width = 12  # 新增“大区”列的列宽
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 13.8
        ws.column_dimensions['N'].width = 13.8
        ws.column_dimensions['O'].width = 13.8

        for row in ws.iter_rows(min_row=1, max_row=max_row):
            ws.row_dimensions[row[0].row].height = 18

        # 定义 J、K、L 列的索引
        jkl_columns = [ord('J') - ord('A') + 1, ord('K') - ord('A') + 1, ord('L') - ord('A') + 1]

        # 遍历每一行，设置样式并进行四舍五入
        for row in range(2, max_row + 1):
            # 检查是否为新增行
            is_new_row = ws.cell(row=row, column=is_new_col_index).value
            if is_new_row:
                for col in range(1, ws.max_column):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                    fill_type="solid")

            # 设置差额列的样式并进行四舍五入
            for col_index in new_col_indices:
                value = ws.cell(row=row, column=col_index).value
                if pd.notna(value):
                    rounded_value = round(value, 2) if abs(value) >= 0.01 else None
                    ws.cell(row=row, column=col_index).value = rounded_value
                    if rounded_value and rounded_value > 0:
                        ws.cell(row=row, column=col_index).font = Font(name='微软雅黑', size=11, bold=True)
                    elif rounded_value and rounded_value < 0:
                        ws.cell(row=row, column=col_index).font = Font(name='微软雅黑', size=11, color="FF0000",
                                                                       bold=True)

            # 处理 J、K、L 列，保留 3 位小数
            for col_index in jkl_columns:
                value = ws.cell(row=row, column=col_index).value
                if pd.notna(value):
                    rounded_value = round(value, 3) if abs(value) >= 0.001 else None
                    ws.cell(row=row, column=col_index).value = rounded_value

        # 删除 is_new 列
        ws.delete_cols(is_new_col_index)

        # 增加汇总行
        summary_row = max_row + 1
        ws.cell(row=summary_row, column=1, value='汇总')
        ws.cell(row=summary_row, column=1).font = bold_font
        for col_index, col_name in enumerate(header, start=1):
            if col_name in compare_columns + new_column_names:
                column_values = [ws.cell(row=row, column=col_index).value for row in range(2, max_row + 1) if
                                 pd.notna(ws.cell(row=row, column=col_index).value)]
                summary_value = sum(column_values)
                rounded_summary = round(summary_value, 2) if abs(summary_value) >= 0.01 else None
                ws.cell(row=summary_row, column=col_index, value=rounded_summary)
                if rounded_summary and rounded_summary > 0:
                    ws.cell(row=summary_row, column=col_index).font = Font(name='微软雅黑', size=11, bold=True)
                elif rounded_summary and rounded_summary < 0:
                    ws.cell(row=summary_row, column=col_index).font = Font(name='微软雅黑', size=11, color="FF0000",
                                                                           bold=True)

        # 冻结首行
        ws.freeze_panes = ws['A2']

        # 设置汇总行样式，明确指定列范围为 A 到 O
        summary_fill = PatternFill(start_color='62B460', end_color='62B460', fill_type='solid')
        for col in range(1, 16):
            ws.cell(row=summary_row, column=col).fill = summary_fill
            ws.cell(row=summary_row, column=col).alignment = center_alignment

    # 保存修改后的 Excel 文件
    wb.save(output_file)
    print(f"比对结果已保存至 {output_file}")
except Exception as e:
    print(f"处理过程中出现错误: {e}")
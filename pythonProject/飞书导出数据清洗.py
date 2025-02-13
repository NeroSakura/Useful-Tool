import os
import re
import requests
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

# 整体功能实现：从飞书提取文档-清洗后输出excel，按八大区分布

# 飞书 API 配置
APP_ID = "cli_a72e948d61d2900e"
APP_SECRET = "DiWuiC8pF7SjZOknpehdvgXWZBmUMann"
ACCESS_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"


def get_desktop_path():
    """获取用户桌面路径"""
    return os.path.join(os.path.expanduser("~"), "Desktop")


def get_access_token():
    """获取访问令牌"""
    payload = {"app_id": APP_ID, "app_secret": APP_SECRET}
    try:
        response = requests.post(ACCESS_TOKEN_URL, json=payload, timeout=10)
        response.raise_for_status()
        return response.json().get("tenant_access_token")
    except Exception as e:
        print(f"❌ 获取Token失败: {str(e)}")
        return None


def parse_document_url(document_url):
    """解析多维表格URL"""
    pattern = r"/base/([a-zA-Z0-9_]+).*table=([a-zA-Z0-9_]+)"
    match = re.search(pattern, document_url)
    if not match:
        return None, None
    return match.group(1), match.group(2)


def get_app_name(access_token, base_token):
    """获取多维表格应用名称"""
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{base_token}"
    headers = {"Authorization": f"Bearer {access_token}"}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        return response.json().get("data", {}).get("name", "未知文档")
    except Exception as e:
        print(f"⚠️ 获取文档名称失败: {str(e)}")
        return "飞书文档"


def get_table_data(access_token, base_token, table_id):
    """获取完整表格数据（自动分页）"""
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{base_token}/tables/{table_id}/records"

    all_records = []
    page_token = ""
    while True:
        params = {"page_size": 100}
        if page_token:
            params["page_token"] = page_token

        try:
            response = requests.get(url, headers=headers, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()

            if data.get("code") != 0:
                print(f"❌ API返回错误: {data.get('msg')}")
                return None

            all_records.extend(data.get("data", {}).get("items", []))

            if not data.get("data", {}).get("has_more"):
                break

            page_token = data.get("data", {}).get("page_token", "")
        except Exception as e:
            print(f"❌ 获取数据失败: {str(e)}")
            return None

    # 打印 API 返回的原始数据
    if all_records:
        print("API 返回的原始数据前5行:")
        for record in all_records[:5]:
            print(record)
    else:
        print("❌ 未获取到任何记录。")

    return {"data": {"items": all_records}}


def save_to_file(df, save_path, file_name, file_format):
    """保存数据到文件"""
    full_path = os.path.join(save_path, f"{file_name}.{file_format}")
    try:
        if file_format == "csv":
            df.to_csv(full_path, index=False, encoding="utf-8-sig")
        elif file_format == "xlsx":
            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                # 写入全国数据到 sheet1
                df.to_excel(writer, sheet_name='全国数据', index=False)
                workbook = writer.book
                worksheet = writer.sheets['全国数据']

                # 自动调整列宽
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.column_dimensions[chr(65 + col_idx)].width = 10  # 修改列宽为10磅

                # 设置行高、字体、水平居中、垂直居中，并框选所有线框
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                        cell.font = openpyxl.styles.Font(name='微软雅黑')
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        worksheet.row_dimensions[cell.row].height = 22.5  # 修改行高为22.5磅

                # 首行标题加粗，并用RGB(153,218,160)填充
                header_row = worksheet[1]
                fill = PatternFill(start_color='99DAA0', end_color='99DAA0', fill_type='solid')
                for cell in header_row:
                    cell.font = Font(name='微软雅黑', bold=True)
                    cell.fill = fill

                regions = ['东南', '中东', '华北', '华南', '西南', '西北', '东北', '华东']
                for region in regions:
                    # 筛选对应大区数据
                    if '大区' in df.columns:
                        region_df = df[df['大区'] == region]
                        # 创建新 sheet 并写入对应大区数据
                        region_sheet = workbook.create_sheet(region)
                        # 写入表头
                        for col_idx, col_name in enumerate(df.columns, start=1):
                            cell = region_sheet.cell(row=1, column=col_idx)
                            cell.value = col_name
                            cell.font = Font(name='微软雅黑', bold=True)
                            cell.fill = fill
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                            cell.border = openpyxl.styles.Border(
                                left=openpyxl.styles.Side(style='thin'),
                                right=openpyxl.styles.Side(style='thin'),
                                top=openpyxl.styles.Side(style='thin'),
                                bottom=openpyxl.styles.Side(style='thin')
                            )
                        # 写入数据
                        for row_idx, row in enumerate(region_df.values, start=2):
                            for col_idx, value in enumerate(row, start=1):
                                cell = region_sheet.cell(row=row_idx, column=col_idx)
                                cell.value = value
                                cell.font = Font(name='微软雅黑')
                                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                                cell.border = openpyxl.styles.Border(
                                    left=openpyxl.styles.Side(style='thin'),
                                    right=openpyxl.styles.Side(style='thin'),
                                    top=openpyxl.styles.Side(style='thin'),
                                    bottom=openpyxl.styles.Side(style='thin')
                                )
                                region_sheet.row_dimensions[cell.row].height = 22.5  # 修改行高为22.5磅

        print(f"✅ 文件已保存至：{full_path}")
        return True
    except Exception as e:
        print(f"❌ 保存文件失败: {str(e)}")
        return False


def process_data(df):
    """处理数据"""
    # 统一字段
    all_fields = set()
    for record in df.to_dict(orient='records'):
        all_fields.update(record.keys())
    new_records = []
    for record in df.to_dict(orient='records'):
        new_record = {}
        for field in all_fields:
            new_record[field] = record.get(field)
        new_records.append(new_record)
    df = pd.DataFrame(new_records)

    # 处理日期格式
    if "日期" in df.columns:
        df["日期"] = pd.to_datetime(df["日期"], unit='ms', utc=True)
        df["日期"] = df["日期"].dt.tz_convert('Asia/Shanghai')
        df["日期"] = df["日期"].dt.strftime('%m月%d日')

    # 处理纵队列，只保留文字
    def extract_text(value):
        if isinstance(value, list):
            if value and isinstance(value[0], dict) and 'text' in value[0]:
                return value[0]['text']
            elif value and isinstance(value[0], str):
                return value[0]
        return value

    if "纵队" in df.columns:
        df["纵队"] = df["纵队"].apply(extract_text)

    # 定义需要保留的列
    columns_to_keep = ["大区", "城市", "纵队", "类别", "执行季度", "执行月", "日期", "客户人数", "预计产出订单金额"]

    # 过滤数据框，只保留指定的列
    df = df.filter(items=columns_to_keep)

    # 按日期列升序排列
    df = df.sort_values(by="日期")
    return df


def get_user_input():
    """获取用户输入的多维表格URL并解析"""
    document_url = input("请输入多维表格URL：").strip()
    if not document_url:
        print("❌ URL不能为空，请重新输入。")
        return None, None
    base_token, table_id = parse_document_url(document_url)
    if not all([base_token, table_id]):
        print("❌ URL格式错误，请确认包含/base/和table参数。")
        return None, None
    return base_token, table_id


def get_file_format_choice():
    """获取用户选择的文件导出格式"""
    print("请选择导出格式：")
    print("1. Excel 文件 (.xlsx)")
    print("2. CSV 文件 (.csv)")
    choice = input("请输入数字选择（默认1）: ").strip() or "1"
    return "xlsx" if choice == "1" else "csv"


def main():
    try:
        # 获取访问令牌
        access_token = get_access_token()
        if not access_token:
            print("❌ 无法获取访问令牌，请检查APP_ID和APP_SECRET是否正确。")
            return

        # 获取用户输入
        base_token, table_id = get_user_input()
        if not all([base_token, table_id]):
            return

        # 获取文档信息
        today = datetime.now().strftime("%Y年%m月%d日")
        file_name = f"智慧中国行{today}数据"

        # 获取表格数据
        print("⏳ 正在获取表格数据...")
        content = get_table_data(access_token, base_token, table_id)
        if not content:
            print("❌ 获取表格数据失败，请检查网络连接或API权限。")
            return

        # 处理数据
        records = content["data"]["items"]
        if not records:
            print("❌ 未获取到有效数据。")
            return
        df = pd.DataFrame([record["fields"] for record in records])
        df = process_data(df)

        # 打印处理后的数据
        print("处理后的数据前5行:")
        print(df.head())

        # 保存文件
        desktop_path = get_desktop_path()
        file_format = get_file_format_choice()
        print(f"Saving file as {file_format} format...")
        if file_format == "xlsx":
            save_to_file(df, desktop_path, file_name, file_format)
        else:
            print("仅支持将大区数据保存到 Excel 文件，请重新选择。")

    except Exception as e:
        print(f"❌ 程序执行过程中发生未知错误: {str(e)}")


if __name__ == "__main__":
    main()
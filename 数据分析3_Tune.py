import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime


def format_sheet(sheet):
    # 设置默认字体和对齐
    font = Font(name='微软雅黑', size=11)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 设置首行格式
    for cell in sheet[1]:
        cell.font = Font(name='微软雅黑', size=11, bold=True)
        cell.alignment = alignment
        cell.border = thin_border
    sheet.row_dimensions[1].height = 32

    # 设置列宽
    for col in sheet.columns:
        column_letter = col[0].column_letter
        sheet.column_dimensions[column_letter].width = 10

    # 设置所有单元格格式
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border

    # 设置数值格式
    format_columns = {
        'SQL $M': numbers.FORMAT_NUMBER_00,
        '订单 $M': numbers.FORMAT_NUMBER_00,
        '商机 $M': numbers.FORMAT_NUMBER_00,
        'SQL达成率': '0.00%',
        '订单转化率': '0.00%',
        '订单达成率': '0.00%',
        '单站产出': '0.00',
        '较上周单站产出': '0.00'
    }

    header = [cell.value for cell in sheet[1]]
    for idx, title in enumerate(header, 1):
        col_letter = chr(64 + idx)
        if title in format_columns:
            for cell in sheet[col_letter][1:]:  # 从第二行开始
                if '达成率' in title or '转化率' in title:
                    if cell.value is not None:
                        cell.number_format = format_columns[title]
                else:
                    cell.number_format = format_columns[title]

    # 设置行高
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 22.5


# 确保以下代码顶格写（不在函数体内）
folder_path = input("请输入文件夹的路径：")
if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
    print("输入的路径不是有效的文件夹。")
else:
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    excel_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith(('.xlsx', '.xlsm'))]

    for file_path in excel_files:
        try:
            wb = load_workbook(file_path)

            for sheet_name in wb.sheetnames:
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # 假设汇总行是最后一行，去除汇总行数据
                if df.apply(lambda row: row.astype(str).str.contains('汇总|总计').any(), axis=1).any():
                    df = df[~df.apply(lambda row: row.astype(str).str.contains('汇总|总计').any(), axis=1)]

                # 将需要汇总的列转换为数值类型
                columns_to_sum = ['SQL $M', '订单 $M', '商机 $M', '高价值客户覆盖数']
                columns_to_average = ['SQL达成率', '订单转化率', '订单达成率']
                for col in columns_to_sum + columns_to_average:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

                # 统计站点数量
                site_counts = df.groupby('大区').size().reset_index(name='站点数量')

                # 以大区为逻辑进行数据汇总
                summary_df = df.groupby('大区', as_index=False).agg({
                    'SQL $M': 'sum',
                    '订单 $M': 'sum',
                    '商机 $M': 'sum',
                    '高价值客户覆盖数': 'sum',
                    'SQL达成率': 'mean',
                    '订单转化率': 'mean',
                    '订单达成率': 'mean'
                })

                # 合并站点数量数据
                summary_df = summary_df.merge(site_counts, on='大区')

                # 计算单站产出
                summary_df['单站产出'] = (summary_df['订单 $M'] / summary_df['站点数量']).round(2)

                # 假设订单差额列名为订单 $M 差额，计算较上周单站产出
                if '订单 $M 差额' in df.columns:
                    order_diff_sum = df.groupby('大区')['订单 $M 差额'].sum().reset_index(name='订单 $M 差额总和')
                    summary_df = summary_df.merge(order_diff_sum, on='大区')
                    summary_df['较上周单站产出'] = (summary_df['订单 $M 差额总和'] / summary_df['站点数量']).round(2)
                else:
                    summary_df['较上周单站产出'] = None

                # 调整列顺序，将站点数量放到第12列
                columns = summary_df.columns.tolist()
                columns.remove('站点数量')
                new_columns_order = columns[:11] + ['站点数量'] + columns[11:]
                summary_df = summary_df[new_columns_order]

                # 新增空列（如果前面计算没涉及到的话）
                new_columns = ['单站产出', '较上周单站产出', '站点数量', '上周站点数量']
                for col in new_columns:
                    if col not in summary_df.columns:
                        summary_df[col] = None

                # 创建新sheet
                new_sheet_name = f'{sheet_name}-区域详情'
                idx = wb.sheetnames.index(sheet_name) + 1
                new_sheet = wb.create_sheet(new_sheet_name, idx)

                # 写入数据
                for r in dataframe_to_rows(summary_df, index=False, header=True):
                    new_sheet.append(r)

                # 应用格式
                format_sheet(new_sheet)

            save_file_path = os.path.join(desktop_path, f'处理后的_{os.path.basename(file_path)}')
            wb.save(save_file_path)
            print(f"文件已保存到: {save_file_path}")

        except Exception as e:
            print(f"处理文件时发生错误: {e}")